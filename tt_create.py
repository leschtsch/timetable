from tkinter import *
import openpyxl
from openpyxl.styles import PatternFill


# this is just key for list.sort() for prior
def priority_sort(a):
    return a[-1]


# this reads list of teachers
def prior():
    global priority
    f = open('данные\\приоритеты.txt', encoding='utf-8')
    priority = [i.split() for i in f]
    priority.sort(key=priority_sort)
    priority = priority[::-1]
    for i in range(len(priority)):
        priority[i][-1] = float(priority[i][-1])
    f.close()


# this is just key for list.sort() for order
def order_sort(a):
    r = 0
    p1 = {str(i): 0 for i in range(-5, 6)}
    p2 = {str(i): 0 for i in range(-5, 6)}
    for i in lessons[a[0][0]][a[0][1]]:
        for z in i:
            p1[z] += 1
    for i in lessons[a[1][0]][a[1][1]]:
        for z in i:
            p2[z] += 1
    for i in lessons[a[0][0]][a[0][1]]:
        for z in i:
            r += int(z) * p1[z] * priority[search(a[0][0])][2]
    for i in lessons[a[1][0]][a[1][1]]:
        for z in i:
            r += int(z) * p2[z] * priority[search(a[1][0])][2]
    return r


# this creates order of pairs
def order():
    global row, after
    f = open('данные\\порядок.txt', encoding='utf-8')
    for i in f:
        a = i.split('\t')
        p1 = a[0].split()
        p2 = a[2].split()
        if a[1] == 'после':
            after.append([[p1[0], (p1[1], p1[2])], [p2[0], (p2[1], p2[2])]])
        elif a[1] == 'подряд':
            row.append([[p1[0], (p1[1], p1[2])], [p2[0], (p2[1], p2[2])]])
        else:
            c.create_text(10, 10, text='неизвестное слово:' + a[1], font='Arial50', fill='#f00000', anchor='nw')
    f.close()
    row.sort(key=order_sort)
    after.sort(key=order_sort)


# this reads teachers' lessons
def create_pairs(a):
    global lessons
    lessons[a] = {}
    f = open('данные\\' + a + '.txt', encoding='utf-8')
    ft = ''.join([i for i in f])
    ft = ft.split('@')
    for i in ft:
        ind = tuple(i.split('#')[0].split('\t'), )
        ind = (ind[0].strip('\n'), ind[1].strip('\n'))
        lessons[a][ind] = i.split('#')[1]
    for i in lessons[a]:
        lessons[a][i] = lessons[a][i].split('\n')
        while lessons[a][i].count(''):
            lessons[a][i].pop(lessons[a][i].index(''))
        for z in range(len(lessons[a][i])):
            lessons[a][i][z] = lessons[a][i][z].split('\t')
    f.close()


# this reads data
def read():
    global timetable, priority, c, lessons
    prior()
    f = open('данные\\приоритеты.txt', 'w', encoding='utf-8')
    for i in priority:
        print('\t'.join(map(str, i)), file=f)
    f.close()
    try:
        for i in priority:
            create_pairs(i[0])
    except FileNotFoundError:
        c.create_text(10, 10, text='не все файлы препов найдены', font='Arial50', fill='#f00000', anchor='nw')
    order()


# this creates order of pairs from excel
def order_excel():
    global row, after, c
    i = 1
    while True:
        if not data['порядок'][i][0].value:
            break
        else:
            if data['порядок'][i][1].value == 'подряд':
                st = data['порядок'][i][0].value.split()
                r = [[st[0], (' '.join(st[1:-1]), st[-1])]]
                st = data['порядок'][i][2].value.split()
                r.append([st[0], (' '.join(st[1:-1]), st[-1])])
                row.append(r)
            elif data['порядок'][i][1].value == 'после':
                st = data['порядок'][i][0].value.split()
                r = [[st[0], (' '.join(st[1:-1]), st[-1])]]
                st = data['порядок'][i][2].value.split()
                r.append([st[0], (' '.join(st[1:-1]), st[-1])])
                after.append(r)
            else:
                c.create_text(10, 10, text='неизвестное слово:' + data['порядок'][i][1].value, font='Arial50',
                              fill='#f00000', anchor='nw')
        i += 1


# this reads list of teachers from excel
def prior_excel():
    global priority
    i = 1
    while True:
        if data['приоритеты'][i][0].value:
            priority.append([data['приоритеты'][i][0].value,
                             data['приоритеты'][i][1].value, data['приоритеты'][i][2].value])
        else:
            break
        i += 1
    priority.sort(key=priority_sort)
    for i in range(len(priority)):
        data['приоритеты'][i + 1][0].value = priority[i][0]
        data['приоритеты'][i + 1][1].value = priority[i][1]
        data['приоритеты'][i + 1][2].value = priority[i][2]


# this reads teachers' lessons from excel
def create_pairs_excel(a):
    global lessons
    lessons[a] = {}
    i = 1
    gtable = []
    for z in range(i, i + 5):
        gtable.append([])
        for q in range(6):
            gtable[z - i].append(str(data[a][z][q].value))
    while True:
        if data[a][i][0].value is None:
            break
        if (data[a][i][3].value is None) and (data[a][i + 1][3].value is None):
            lessons[a][(data[a][i][0].value, str(data[a][i][1].value))] = [[x for x in q] for q in gtable]
        elif (data[a][i][3].value is None) and (data[a][i + 1][3].value is not None):
            lessons[a][(data[a][i][0].value, str(data[a][i][1].value))] = []
            for z in range(i + 1, i + 6):
                lessons[a][(data[a][i][0].value, str(data[a][i][1].value))].append([])
                for q in range(6):
                    lessons[a][(data[a][i][0].value, str(data[a][i][1].value))][z - i - 1].append(
                        str(data[a][z][q].value))
        i += 1


# this reads data from excel
def read_excel():
    prior_excel()
    try:
        for i in priority:
            create_pairs_excel(i[0])
    except KeyError:
        c.create_text(10, 10, text='не все листы препов найдены', font='Arial50', fill='#f00000', anchor='nw')
    order_excel()


# this evaluates the timetable
def evaluate(ttl):
    summa = 0
    for i in ttl:
        for z in range(len(i)):
            for q in range(len(i[z])):
                if i[z][q]:
                    summa += priority[search(i[z][q][0])][2] * int(lessons[i[z][q][0]][i[z][q][1]][z][q])
    return int(summa)


# this searches teacher in priority
def search(name):
    global priority
    for i in range(len(priority)):
        if priority[i][0] == name:
            return i


# this is sort for orders
def order_index_sort(a, n1, n2):
    global lessons
    t1 = lessons[n1[0]][n1[1]]
    t2 = lessons[n2[0]][n2[1]]
    p1 = priority[search(n1[0])][2]
    p2 = priority[search(n2[0])][2]
    s1 = [[i] for i in a.copy()]
    s2 = []
    while len(s1) > 1:  # merge sort
        while len(s1) > 1:
            n = []
            while True:
                if int(t1[s1[0][0][0]][s1[0][0][1]]) * p1 + int(t2[s1[0][0][2]][s1[0][0][3]]) * p2 < int(
                        t1[s1[1][0][0]][s1[1][0][1]]) * p1 + int(t2[s1[1][0][2]][s1[1][0][3]]) * p2:
                    n.append(s1[0].pop(0))
                else:
                    n.append(s1[1].pop(0))
                if not len(s1[0]):
                    n.extend(s1[1])
                    break
                if not len(s1[1]):
                    n.extend(s1[0])
                    break
            s2.append(n)
            s1.pop(0)
            s1.pop(0)
        if len(s1):
            s2.append(s1[0])
        s1 = s2.copy()
        s2 = []
    return s1[0]


# this searches most wanted pair of lessons in a row
def row_search_max(a, ttbl):
    m = []
    for i in range(4):
        for z in range(6):
            if not (ttbl[int(a[0][1][1]) - 6][i][z] or ttbl[int(a[1][1][1]) - 6][i + 1][z]):
                m.append([i, z, i + 1, z])
    m = order_index_sort(m, a[0], a[1])
    if len(m) > lengen:
        m = m[-lengen:]
    return m


# this searches most wanted pair of lessons
def after_search_max(a, ttbl):
    m = []
    for i in range(len(lessons[a[0][0]][a[0][1]])):
        for z in range(len(lessons[a[0][0]][a[0][1]][i])):
            for q in range(len(lessons[a[1][0]][a[1][1]])):
                for r in range(len(lessons[a[1][0]][a[1][1]][q])):
                    if q * len(lessons[a[1][0]][a[1][1]][q]) + r > i * len(lessons[a[0][0]][a[0][1]][i]) + z:
                        if not (ttbl[int(a[0][1][1]) - 6][i][z] or ttbl[int(a[1][1][1]) - 6][q][r]):
                            m.append([i, z, q, r])
                    m = order_index_sort(m, a[0], a[1])
                    if len(m) > lengen:
                        m = m[-lengen:]
    return m


# this is copy for ttable from addd
def tt_copy(ttable):
    r = []
    for i in range(len(ttable)):
        r.append([])
        for z in range(len(ttable[i])):
            r[i].append(ttable[i][z].copy())
    return r


# this is a part of generate
def addd(a, i, z):
    global lessons
    res = []
    m = '5'
    flag = 0
    full = False
    for ii in a:
        if full:
            break
        for zz in ii:
            if full:
                break
            for q in zz:
                if full:
                    break
                if not q:
                    full = True
                if full:
                    break
    if full:
        while flag < lengen:
            if int(m) < -5:
                return []
            for q in range(len(lessons[i][z])):
                for x in range(len(lessons[i][z][q])):
                    if (lessons[i][z][q][x] == m) and (not a[int(z[1]) - 6][q][x]):
                        if ((not a[0][q][x]) or (a[0][q][x][0] != i)) and (
                                (not a[1][q][x]) or (a[1][q][x][0] != i)) and (
                                (not a[2][q][x]) or (a[2][q][x][0] != i)) and (
                                (not a[3][q][x]) or (a[3][q][x][0] != i)) and (
                                (not a[4][q][x]) or (a[4][q][x][0] != i)) and (
                                (not a[5][q][x]) or (a[5][q][x][0] != i)):
                            a2 = tt_copy(a)
                            a2[int(z[1]) - 6][q][x] = [i, z]
                            res.append(a2)
                            flag += 1
            m = str(int(m) - 1)
    if flag == 0:
        return []
    return res


# this is addd for pair of lessons
def addd_2(a, option, ttbl):
    if option == 'row':
        m = row_search_max(a, ttbl)
    elif option == 'after':
        m = after_search_max(a, ttbl)
    if not m:
        return []
    res = []
    for i in m:
        ttbln = tt_copy(ttbl)
        ttbln[int(a[0][1][1]) - 6][i[0]][i[1]] = a[0]
        ttbln[int(a[1][1][1]) - 6][i[2]][i[3]] = a[1]
        res.append(ttbln)
    return res


# this is list.copy() for lessons
def lessons_copy(lessons):
    r = {}
    for i in lessons:
        r[i] = {}
        for z in lessons[i]:
            r[i][z] = []
            for q in lessons[i][z]:
                r[i][z].append(q.copy())
    return r


# this is just key for list.sort() for generate
def pairs_sort(a):
    r = 0
    p1 = {str(i): 0 for i in range(-5, 6)}
    for i in lessons[a[0]][a[1]]:
        for z in i:
            p1[z] += 1
    for i in lessons[a[0]][a[1]]:
        for z in i:
            r += int(z) * p1[z] * priority[search(a[0])][2]
    return r


# this uses normal generation
def generate():
    global lessons, row, after
    ttable = [[[[] for z in range(6)] for i in range(5)] for j in range(6)]
    ttables = [ttable]
    pairs = []
    for i in lessons:
        for z in lessons[i]:
            pairs.append([i, z])
    pairs.sort(key=pairs_sort)

    while row:  # this inserts row
        ttables2 = []
        for i in ttables:
            n = addd_2(row[0], 'row', i)
            if n:
                ttables2.extend(n)
            else:
                return []
        pairs.remove([row[0][0][0], row[0][0][1]])
        pairs.remove([row[0][1][0], row[0][1][1]])
        row.pop(0)
        ttables = ttables2
        ttables.sort(key=evaluate)
        if len(ttables) > lengen:
            ttables = ttables[-lengen:]

    while after:  # this inserts after
        ttables2 = []
        for i in ttables:
            n = addd_2(after[0], 'after', i)
            if n:
                ttables2.extend(n)
            else:
                return []
        pairs.remove([row[0][0][0], row[0][0][1]])
        pairs.remove([row[0][1][0], row[0][1][1]])
        after.pop(0)
        ttables = ttables2
        ttables.sort(key=evaluate)
        if len(ttables) > lengen:
            ttables = ttables[-lengen:]

    while pairs:  # this inserts others
        ttables2 = []
        for q in ttables:
            n = addd(q, pairs[0][0], pairs[0][1])
            if n:
                ttables2.extend(n)
            else:
                return []
        ttables = ttables2
        pairs.pop(0)
        ttables.sort(key=evaluate)
        if len(ttables) > lengen:
            ttables = ttables[-lengen:]
    return ttables


# this creates vertical text
def create_vertical_text(a, b, tex):
    d = 0
    for i in tex:
        c.create_text(a, b + d, text=i, font='Arial15', fill=tcolor)
        d += 17


# this draws timetable
def draw(ttl, conven):
    global c, timetable
    c.delete('all')
    w = 180
    for i in range(len(ttl)):
        for z in range(len(ttl[i])):
            for q in range(len(ttl[i][z])):
                if not ttl[i][z][q]:
                    c.create_rectangle(30 + w * i, 60 + 30 * q + 50 * q * 5 + 50 * z, 30 + w * (i + 1),
                                       60 + 30 * q + 50 * q * 5 + 50 * (z + 1), fill='#dddddd')
                else:
                    color = priority[search(ttl[i][z][q][0])][1]
                    c.create_rectangle(30 + w * i, 60 + 30 * q + 50 * q * 5 + 50 * z, 30 + w * (i + 1),
                                       60 + 30 * q + 50 * q * 5 + 50 * (z + 1),
                                       fill=color)
                    center = ((60 + w * (2 * i + 1)) // 2, (100 * q * 5 + 50 * z + 120 + 60 * q + 50 * (z + 1)) // 2)
                    tx = ttl[i][z][q][1][0] + '       ' + ttl[i][z][q][0]
                    tx.strip('\n')
                    c.create_text(center[0], center[1], text=tx, font='Arial 10', )

    c.create_text(900, 1750, text='сгенерировано ' + str(len(ttbl)) + ' расписаний из ' + str(lengen), font='Arial15',
                  fill=tcolor)
    c.create_text(200, 1750, text='текущее расписание: ' + str(cttl + 1) + ' из ' + str(len(ttbl)), font='Arial15',
                  fill=tcolor)
    c.create_text(200, 1800, text='удобство текущего расписания: ' + str(conven), font='Arial15', fill=tcolor)
    create_vertical_text(15, 70, 'Понедельник')
    create_vertical_text(15, 350, 'Вторник')
    create_vertical_text(15, 630, 'Среда')
    create_vertical_text(15, 910, 'Четверг')
    create_vertical_text(15, 1190, 'Пятница')
    create_vertical_text(15, 1470, 'Суббота')

    for i in range(6):
        for z in range(6):
            c.create_text(50 + i * 180, 50 + z * 280, text=str(i + 6), font='Arial15', fill=tcolor)


# this is def for 'forward' button
def button_forward():
    global cttl, ttbl, convenience
    if ttbl:
        if cttl < len(ttbl) - 1:
            cttl += 1
            convenience = evaluate(ttbl[cttl])
            draw(ttbl[cttl], convenience)
    else:
        c.create_text(175, 10, text='не удалось составить расписание', fill=tcolor)


# this is def for 'back' button
def button_back():
    global cttl, ttbl, convenience
    if ttbl:
        if cttl > 0:
            cttl -= 1
            convenience = evaluate(ttbl[cttl])
            draw(ttbl[cttl], convenience)
    else:
        c.create_text(175, 10, text='не удалось составить расписание', fill=tcolor)


# this saves tiumetables in excel
def save():
    global ttbl, c
    try:
        wb = openpyxl.Workbook()
        for i in range(len(ttbl)):
            wb.create_sheet(str(i + 1))
            for z in range(len(ttbl[i])):
                for q in range(len(ttbl[i][z])):
                    for x in range(len(ttbl[i][z][q])):
                        if ttbl[i][z][q][x]:
                            wb[str(i + 1)].cell(row=3 + x * 6 + q, column=z + 2).value = ttbl[i][z][q][x][
                                                                                             0] + ' ' + ' '.join(
                                ttbl[i][z][q][x][1])
                            wb[str(i + 1)].cell(row=3 + x * 6 + q, column=z + 2).fill = PatternFill(fill_type='solid',
                                                                                                    start_color='FF' +
                                                                                                                priority[
                                                                                                                    search(
                                                                                                                        ttbl[
                                                                                                                            i][
                                                                                                                            z][
                                                                                                                            q][
                                                                                                                            x][
                                                                                                                            0])][
                                                                                                                    1][
                                                                                                                1:].upper(),
                                                                                                    end_color='FF' +
                                                                                                              priority[
                                                                                                                  search(
                                                                                                                      ttbl[
                                                                                                                          i][
                                                                                                                          z][
                                                                                                                          q][
                                                                                                                          x][
                                                                                                                          0])][
                                                                                                                  1][
                                                                                                              1:].upper())
                        else:
                            wb[str(i + 1)].cell(row=3 + x * 6 + q, column=z + 2).value = 'пусто'
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        wb.save('timetables.xlsx')
    except PermissionError:
        c.delete('all')
        c.create_text(10, 10, text='файл открыт другой программой', font='Arial50', fill='#f00000', anchor='nw')


# this reads configuration
def config():
    global lengen, lenshow, bgcolor, btcolor, tcolor
    f = open('config.txt')
    conf = [i.strip() for i in f]
    lengen = int(conf[0])
    lenshow = int(conf[1])
    bgcolor = conf[2]
    btcolor = conf[3]
    tcolor = conf[4]
    if lengen > 20:
        lengen = 20
    if lenshow > 20:
        lenshow = 20
    if lenshow < 1:
        lenshow = 1
    if lengen < 1:
        lengen = 1
    if lengen < lenshow:
        lengen = lenshow
    conf[0] = lengen
    conf[1] = lenshow
    f.close()
    f = open('config.txt', 'w')
    for i in conf:
        print(str(i), file=f)
    f.close()


bgcolor = ''
btcolor = ''
tcolor = ''
lengen = 0
lenshow = 0
config()
timetable = Toplevel()
timetable['bg'] = bgcolor
timetable.title('Расписание')
timetable.geometry('1250x600+10+30')
bb = Button(timetable, text='Назад', bg=btcolor, command=button_back, fg=tcolor)
bb.pack(side='left')
frame = Frame(timetable, width=1160, height=700, bg=bgcolor)
frame.pack(side='left')
c = Canvas(frame, width=1140, height=1900, bg=bgcolor, scrollregion=(0, 0, 1000, 1900))
s = Scrollbar(frame)
c.config(yscrollcommand=s.set)
s.config(command=c.yview)
c.pack(side='left')
s.pack(fill='y', expand=True, side='right')
bf = Button(timetable, text='вперед', bg=btcolor, command=button_forward, fg=tcolor)
bf.pack(side='left')
convenience = 0
lessons = {}
after = []
row = []
priority = []
try:
    data = openpyxl.open('данные.xlsx')
    read_excel()
    data.save('данные.xlsx')
    ttbl = generate()
except PermissionError:
    c.create_text(10, 10, text='файл открыт другой программой', font='Arial50', fill='#f00000', anchor='nw')
ttbl = ttbl[-lenshow:]
ttbl = ttbl[::-1]
cttl = 0
if ttbl:
    save()
    convenience = evaluate(ttbl[cttl])
    draw(ttbl[cttl], convenience)
else:
    c.create_text(175, 10, text='не удалось составить расписание', fill=tcolor)
